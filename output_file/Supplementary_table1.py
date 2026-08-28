#!/usr/bin/env python3
"""Supplementary Table 1: dataset details for the three adopted cohorts.

One sheet per dataset (from checkpoint ``obs.csv``):
  - GSE155622
  - GSE141259
  - HGSOC

Default output:
  output_file/Supplementary_table1.xlsx

Usage:
  python output_file/Supplementary_table1.py
  python output_file/Supplementary_table1.py /path/to/out.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

CK_PAIN = ROOT / (
    "GSE155622_checkpoints_3000_3000_384_0.05_recon0.01_lossnorm_qp_d0p01_z0p5_k0p2_ld1"
)
CK_LUNG = ROOT / (
    "GSE141259_checkpoints_5000_5000_512_0.01_recon0.1_valD28_timeX_lossnorm_qp_d0p01_z0p1_k0p2"
)
CK_HG = ROOT / (
    "HGSOC_checkpoints_3000_3000_512_0.05_recon0.01_nactpair_lossnorm_qp_d0p01_z0p5_k0p2_ld1"
)
LUNG_MAP = CK_LUNG / "figures" / "GSE141259_metacelltype_formal_label_mapping.csv"

DEFAULT_OUT = Path(__file__).resolve().parent / "Supplementary_table1.xlsx"

PAIN_STAGES = ["Control", "SNI 6h", "SNI 24h", "SNI 2d", "SNI 7d", "SNI 14d"]
LUNG_STAGES = ["D0", "D3", "D7", "D10", "D14", "D21", "D28"]
HG_TYPES = ["EOC", "Immune", "Stromal"]
HG_PHASES = ["treatment-naive", "post-NACT"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
META_KEY_FONT = Font(bold=True, size=10)


def _autosize(ws, min_width: float = 10, max_width: float = 42) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min_width
        for cell in col:
            if cell.value is None:
                continue
            width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def _write_section(ws, row: int, title: str) -> int:
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    return row + 1


def _write_kv(ws, row: int, items: list[tuple[str, object]]) -> int:
    for k, v in items:
        c0 = ws.cell(row=row, column=1, value=k)
        c0.font = META_KEY_FONT
        ws.cell(row=row, column=2, value=v)
        row += 1
    return row


def _write_df(ws, row: int, df: pd.DataFrame, *, index: bool = False) -> int:
    out = df.reset_index() if index else df
    for r_i, rec in enumerate(dataframe_to_rows(out, index=False, header=True)):
        for c_i, val in enumerate(rec, start=1):
            cell = ws.cell(row=row + r_i, column=c_i, value=val)
            if r_i == 0:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(wrap_text=True, vertical="center")
    return row + len(out) + 1 + 1  # header + data + blank


def _pain_tables() -> tuple[list[tuple[str, object]], pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    obs = pd.read_csv(CK_PAIN / "obs.csv", usecols=["condition", "annotation"], low_memory=False)
    types = (
        obs["annotation"]
        .astype(str)
        .value_counts()
        .reindex(
            ["Satellite", "Neuron", "Schwann", "Immune", "VEC", "VECC", "Fibroblast", "VSMC", "RBC"]
        )
        .dropna()
        .index.tolist()
    )
    # keep any leftover types
    for t in obs["annotation"].astype(str).value_counts().index:
        if t not in types:
            types.append(t)

    stage_n = (
        obs["condition"]
        .astype(str)
        .value_counts()
        .reindex(PAIN_STAGES)
        .fillna(0)
        .astype(int)
        .rename("n_cells")
        .rename_axis("stage")
        .reset_index()
    )
    type_n = (
        obs["annotation"]
        .astype(str)
        .value_counts()
        .reindex(types)
        .fillna(0)
        .astype(int)
        .rename("n_cells")
        .rename_axis("cell_type")
        .reset_index()
    )
    type_n["fraction"] = (type_n["n_cells"] / type_n["n_cells"].sum()).round(4)

    cross = (
        obs.groupby([obs["condition"].astype(str), obs["annotation"].astype(str)])
        .size()
        .unstack(fill_value=0)
        .reindex(index=PAIN_STAGES, columns=types, fill_value=0)
    )
    cross.insert(0, "n_total", cross.sum(axis=1))
    cross = cross.reset_index().rename(columns={"condition": "stage"})

    meta = [
        ("Dataset ID", "GSE155622"),
        ("Display name", "Neuropathic Pain (SNI)"),
        ("Biology", "Mouse DRG after spared nerve injury (SNI); injury-time course"),
        ("Species / tissue", "Mus musculus / dorsal root ganglion (DRG)"),
        ("Temporal axis", "condition: Control → SNI 6h → 24h → 2d → 7d → 14d (6 stages)"),
        ("n_cells (adopted checkpoint)", int(len(obs))),
        ("n_stages", 6),
        ("n_cell_types (annotation)", int(obs["annotation"].nunique())),
        ("Cell-type key", "annotation (9 major compartments)"),
        ("Adopted checkpoint", CK_PAIN.name),
        ("Training panel / latent dim", "3000 genes / dim 384"),
        ("Validation mode", "random cell holdout"),
    ]
    return meta, stage_n, type_n, cross


def _lung_tables() -> tuple[list[tuple[str, object]], pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    obs = pd.read_csv(
        CK_LUNG / "obs.csv",
        usecols=["stage", "annotation", "orig.ident"],
        low_memory=False,
    )
    map_df = pd.read_csv(LUNG_MAP)
    formal = dict(zip(map_df["metacelltype"].astype(str), map_df["formal_label"].astype(str)))
    types = [t for t in map_df["metacelltype"].astype(str).tolist() if t in set(obs["annotation"].astype(str))]

    stage_n = (
        obs.groupby(obs["stage"].astype(str))
        .agg(n_cells=("annotation", "size"), n_mice=("orig.ident", "nunique"))
        .reindex(LUNG_STAGES)
        .fillna(0)
        .astype(int)
        .reset_index()
        .rename(columns={"stage": "stage"})
    )
    type_n = (
        obs["annotation"]
        .astype(str)
        .value_counts()
        .reindex(types)
        .fillna(0)
        .astype(int)
        .rename("n_cells")
        .rename_axis("cell_type")
        .reset_index()
    )
    type_n.insert(1, "formal_label", type_n["cell_type"].map(formal))
    type_n["fraction"] = (type_n["n_cells"] / type_n["n_cells"].sum()).round(4)

    cross = (
        obs.groupby([obs["stage"].astype(str), obs["annotation"].astype(str)])
        .size()
        .unstack(fill_value=0)
        .reindex(index=LUNG_STAGES, columns=types, fill_value=0)
    )
    # rename columns to formal labels for readability
    cross = cross.rename(columns=formal)
    cross.insert(0, "n_total", cross.sum(axis=1))
    cross = cross.reset_index().rename(columns={"stage": "stage"})

    meta = [
        ("Dataset ID", "GSE141259"),
        ("Display name", "Bleomycin Lung Injury"),
        ("Biology", "Mouse whole-lung after bleomycin; alveolar / immune remodeling over time"),
        ("Species / tissue", "Mus musculus / whole lung"),
        ("Temporal axis", "stage: D0, D3, D7, D10, D14, D21, D28 (7 stages)"),
        ("n_cells (adopted checkpoint)", int(len(obs))),
        ("n_stages", 7),
        ("n_mice (orig.ident)", int(obs["orig.ident"].nunique())),
        ("n_major_cell_types (annotation / metacelltype)", int(obs["annotation"].nunique())),
        ("Cell-type key", "annotation (15 major metacelltypes)"),
        ("Adopted checkpoint", CK_LUNG.name),
        ("Training panel / latent dim", "5000 genes / dim 512"),
        ("Validation mode", "D28 time extrapolation (valholdD28)"),
    ]
    return meta, stage_n, type_n, cross


def _hgsoc_tables() -> tuple[
    list[tuple[str, object]],
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
]:
    obs = pd.read_csv(
        CK_HG / "obs.csv",
        usecols=[
            "patient_id",
            "treatment_phase",
            "annotation",
            "stage",
            "anatomical_location",
            "sample",
        ],
        low_memory=False,
    )
    obs = obs[obs["annotation"].isin(HG_TYPES)].copy()

    def _pid_key(p: str) -> int:
        return int("".join(ch for ch in str(p) if ch.isdigit()) or 0)

    patients = sorted(obs["patient_id"].astype(str).unique(), key=_pid_key)

    # patient × phase totals + clinical stage
    rows = []
    for pid in patients:
        clin = obs.loc[obs.patient_id == pid, "stage"].astype(str).mode().iloc[0]
        for ph in HG_PHASES:
            sub = obs[(obs.patient_id == pid) & (obs.treatment_phase == ph)]
            row = {
                "patient_id": pid,
                "clinical_stage": clin,
                "treatment_phase": ph,
                "phase_short": "TN" if ph == "treatment-naive" else "PN",
                "n_total": int(len(sub)),
            }
            for t in HG_TYPES:
                row[f"n_{t}"] = int((sub.annotation == t).sum())
            rows.append(row)
    patient_phase = pd.DataFrame(rows)

    type_n = (
        obs["annotation"]
        .astype(str)
        .value_counts()
        .reindex(HG_TYPES)
        .fillna(0)
        .astype(int)
        .rename("n_cells")
        .rename_axis("cell_type")
        .reset_index()
    )
    type_n["fraction"] = (type_n["n_cells"] / type_n["n_cells"].sum()).round(4)

    phase_n = (
        obs["treatment_phase"]
        .astype(str)
        .value_counts()
        .reindex(HG_PHASES)
        .fillna(0)
        .astype(int)
        .rename("n_cells")
        .rename_axis("treatment_phase")
        .reset_index()
    )
    phase_n["phase_short"] = phase_n["treatment_phase"].map(
        {"treatment-naive": "TN", "post-NACT": "PN"}
    )

    samples = (
        obs.groupby(
            [
                obs["patient_id"].astype(str),
                obs["treatment_phase"].astype(str),
                obs["sample"].astype(str),
                obs["anatomical_location"].astype(str),
                obs["stage"].astype(str),
            ],
            observed=True,
        )
        .size()
        .reset_index(name="n_cells")
        .rename(
            columns={
                "patient_id": "patient_id",
                "treatment_phase": "treatment_phase",
                "sample": "sample",
                "anatomical_location": "anatomical_location",
                "stage": "clinical_stage",
            }
        )
    )
    samples["patient_id"] = pd.Categorical(samples["patient_id"], categories=patients, ordered=True)
    samples["treatment_phase"] = pd.Categorical(
        samples["treatment_phase"], categories=HG_PHASES, ordered=True
    )
    samples = samples.sort_values(["patient_id", "treatment_phase"]).reset_index(drop=True)

    meta = [
        ("Dataset ID", "HGSOC"),
        ("Display name", "HGSOC (NACT-paired)"),
        ("Biology", "High-grade serous ovarian cancer; paired treatment-naive vs post-NACT samples"),
        ("Species / tissue", "Homo sapiens / peritoneal / omental / related sites"),
        ("Temporal / pairing axis", "treatment_phase: treatment-naive (TN) → post-NACT (PN)"),
        ("n_cells (adopted checkpoint)", int(len(obs))),
        ("n_patients (paired)", int(len(patients))),
        ("n_samples", int(obs["sample"].nunique())),
        ("n_cell_types (annotation)", 3),
        ("Cell-type key", "annotation (EOC / Immune / Stromal)"),
        ("Clinical stages present", ", ".join(sorted(obs["stage"].astype(str).unique()))),
        ("Adopted checkpoint", CK_HG.name),
        ("Training panel / latent dim", "3000 genes / dim 512"),
        ("Validation mode", "patient holdout (nactpair)"),
    ]
    return meta, phase_n, type_n, patient_phase, samples


def _build_sheet_pain(wb: Workbook) -> None:
    ws = wb.create_sheet("GSE155622", 0)
    meta, stage_n, type_n, cross = _pain_tables()
    r = 1
    r = _write_section(ws, r, "A. Overview")
    r = _write_kv(ws, r, meta) + 1
    r = _write_section(ws, r, "B. Cells per injury stage")
    r = _write_df(ws, r, stage_n) + 1
    r = _write_section(ws, r, "C. Cells per cell type (overall)")
    r = _write_df(ws, r, type_n) + 1
    r = _write_section(ws, r, "D. Stage × cell-type counts")
    _write_df(ws, r, cross)
    _autosize(ws)


def _build_sheet_lung(wb: Workbook) -> None:
    ws = wb.create_sheet("GSE141259", 1)
    meta, stage_n, type_n, cross = _lung_tables()
    r = 1
    r = _write_section(ws, r, "A. Overview")
    r = _write_kv(ws, r, meta) + 1
    r = _write_section(ws, r, "B. Cells per time point (and n mice)")
    r = _write_df(ws, r, stage_n) + 1
    r = _write_section(ws, r, "C. Cells per major cell type (overall)")
    r = _write_df(ws, r, type_n) + 1
    r = _write_section(ws, r, "D. Stage × major cell-type counts")
    _write_df(ws, r, cross)
    _autosize(ws, max_width=28)


def _build_sheet_hgsoc(wb: Workbook) -> None:
    ws = wb.create_sheet("HGSOC", 2)
    meta, phase_n, type_n, patient_phase, samples = _hgsoc_tables()
    r = 1
    r = _write_section(ws, r, "A. Overview")
    r = _write_kv(ws, r, meta) + 1
    r = _write_section(ws, r, "B. Cells per treatment phase")
    r = _write_df(ws, r, phase_n) + 1
    r = _write_section(ws, r, "C. Cells per cell type (overall)")
    r = _write_df(ws, r, type_n) + 1
    r = _write_section(ws, r, "D. Per-patient TN / PN cell counts by type")
    r = _write_df(ws, r, patient_phase) + 1
    r = _write_section(ws, r, "E. Sample inventory (patient × phase × site)")
    _write_df(ws, r, samples)
    _autosize(ws, max_width=36)


def compose(out: Path | None = None) -> Path:
    out = Path(out or DEFAULT_OUT)
    wb = Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)
    _build_sheet_pain(wb)
    _build_sheet_lung(wb)
    _build_sheet_hgsoc(wb)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}", flush=True)
    return out


def main(argv: list[str] | None = None) -> int:
    argv = list(sys.argv[1:] if argv is None else argv)
    out = Path(argv[0]) if argv else DEFAULT_OUT
    compose(out=out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
